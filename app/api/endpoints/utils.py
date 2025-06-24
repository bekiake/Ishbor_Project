"""
Utilities API endpointlari

Turli yordamchi API endpointlari
"""
from typing import Any, List, Dict

from fastapi import APIRouter, Depends, HTTPException, status, Query, UploadFile, File
from fastapi.responses import FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
import io
import csv
import json
import os
from pathlib import Path
import openpyxl
from io import BytesIO
from fastapi import Response, HTTPException
from app.database import get_db as get_async_db 
from app.core.security import get_current_active_user
from app.core.settings import settings
from app.models.models import User
from app.models import models
from app.crud import worker as worker_crud
from app.crud import feedback as feedback_crud
from app.crud import user as user_crud

router = APIRouter()




@router.get("/stats/system")
async def get_system_stats(
    db: AsyncSession = Depends(get_async_db),
    current_user: User = Depends(get_current_active_user),
) -> dict[str, Any]:
    """
    Tizimning umumiy statistikalarini olish:
      - foydalanuvchilar soni
      - ishchilar soni
      - fikrlar statistikasi
      - top ko‘nikmalar va tillar
    """
    # Foydalanuvchilar soni
    users_count = await user_crud.get_user_count(db)
    active_users_count = await user_crud.get_active_user_count(db)

    # Ishchilar statistikasi
    worker_stats = await worker_crud.get_worker_statistics(db) or {}

    # Fikrlar statistikasi
    feedback_stats = await feedback_crud.get_feedback_statistics(db) or {}

    return {
        "total_users": users_count,
        "active_users": active_users_count,
        "total_workers": worker_stats.get("total_workers", 0),
        "active_workers": worker_stats.get("active_workers", 0),
        "total_feedbacks": feedback_stats.get("total_feedbacks", 0),
        "active_feedbacks": feedback_stats.get("active_feedbacks", 0),
        "average_rating": feedback_stats.get("average_rating", 0),
        "rating_distribution": feedback_stats.get("rating_distribution", {}),
        "payment_distribution": worker_stats.get("payment_distribution", {}),
        "gender_distribution": worker_stats.get("gender_distribution", {}),
        "disability_distribution": worker_stats.get("disability_distribution", {}),
        "top_skills": [],       # Agar kelajakda qo‘shsang, shu yerga yoz
        "top_languages": [],    # Shu yerga ham
    }



@router.get("/export/workers")
async def export_workers_excel(
        db: AsyncSession = Depends(get_async_db)
) -> Response:
    """
    Ishchilar ro'yxatini Excel formatida eksport qilish (ommaviy endpoint)
    """
    try:
        # Barcha ishchilarni olish
        stmt = select(models.Worker)
        result = await db.execute(stmt)
        workers = result.scalars().all()

        # Excel faylini xotiraga yaratish (disk o'rniga)
        output = BytesIO()
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Ishchilar"

        # Ustun nomlari
        headers = [
            "ID", "Telegram ID", "Ism", "Haqida", "Yosh", "Telefon", "Jinsi",
            "To'lov turi", "Kunlik to'lov", "Tillar", "Ko'nikmalar",
            "Manzil", "Nogironlik darajasi", "Rasm", "Yaratilgan sana", "Yangilangan sana", "Faol"
        ]

        # Ustun nomlarini yozish
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = openpyxl.styles.Font(bold=True)
            # Ustun kengligini o'rnatish
            worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 15

        # Ma'lumotlarni yozish
        for row_num, worker in enumerate(workers, 2):
            try:
                worksheet.cell(row=row_num, column=1).value = worker.id
                worksheet.cell(row=row_num, column=2).value = worker.telegram_id
                worksheet.cell(row=row_num, column=3).value = worker.name
                worksheet.cell(row=row_num, column=4).value = worker.about
                worksheet.cell(row=row_num, column=5).value = worker.age
                worksheet.cell(row=row_num, column=6).value = worker.phone
                worksheet.cell(row=row_num, column=7).value = worker.gender
                worksheet.cell(row=row_num, column=8).value = worker.payment_type
                worksheet.cell(row=row_num, column=9).value = worker.daily_payment
                worksheet.cell(row=row_num, column=10).value = worker.languages
                worksheet.cell(row=row_num, column=11).value = worker.skills
                worksheet.cell(row=row_num, column=12).value = worker.location
                worksheet.cell(row=row_num, column=13).value = worker.disability_degree  # Yangi maydon
                worksheet.cell(row=row_num, column=14).value = worker.image
                worksheet.cell(row=row_num, column=15).value = worker.created_at.isoformat() if worker.created_at else ""
                worksheet.cell(row=row_num, column=16).value = worker.updated_at.isoformat() if worker.updated_at else ""
                worksheet.cell(row=row_num, column=17).value = "Ha" if worker.is_active else "Yo'q"
            except Exception as e:
                # Qatorlardagi alohida xatolar butun jarayonni to'xtatmasligi kerak
                continue

        # Faylni BytesIO obyektiga saqlash
        workbook.save(output)
        output.seek(0)  # Pozitsiyasini boshiga qaytarish

        # Response qaytarish
        headers = {
            'Content-Disposition': 'attachment; filename="workers_export.xlsx"',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
        return Response(content=output.getvalue(), headers=headers)

    except Exception as e:
        # Xatoni log qilish
        import logging
        logging.error(f"Export workers error: {str(e)}")

        # Xato haqida xabar qaytarish
        raise HTTPException(status_code=500, detail=f"Export yaratishda xatolik: {str(e)}")


@router.get("/skills", response_model=List[str])
async def get_all_skills(db: AsyncSession = Depends(get_async_db)):
    return await worker_crud.get_all_skill_names(db)

# @router.get("/disability_degrees", response_model=List[str])
# async def get_all_disability_degrees(db: AsyncSession = Depends(get_async_db)):
#     """Barcha nogironlik darajalarini qaytarish"""
#     return await worker_crud.get_all_disability_degrees(db)